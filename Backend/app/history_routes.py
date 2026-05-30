"""
history_routes.py
Va en: Backend/app/history_routes.py
Usa la tabla 'positions' que ya existe en tu BD.
"""
from __future__ import annotations
from fastapi import FastAPI, Query, HTTPException
from app.db import get_conn
from datetime import datetime, timedelta
import math


def register_history_routes(app: FastAPI) -> None:

    # ── GET /app/history/route ────────────────────────────────
    @app.get("/app/history/route")
    def get_route(
        vehicle_id: str,
        date: str = Query(..., description="YYYY-MM-DD"),
        client_id: str = Query(...),
    ):
        # Verificar que el vehículo pertenece al cliente
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute("""
                SELECT v.id FROM vehicles v
                JOIN app_clients c ON c.vehicle_id = v.id
                WHERE v.id = %s AND c.id = %s
            """, (vehicle_id, client_id))
            if not cur.fetchone():
                raise HTTPException(status_code=403, detail="Vehículo no encontrado")

            cur.execute("""
                SELECT lat, lng, speed, heading, event_type, created_at
                FROM positions
                WHERE vehicle_id = %s
                  AND DATE(created_at::timestamp) = %s::date
                ORDER BY created_at ASC
            """, (vehicle_id, date))
            points = [dict(r) for r in cur.fetchall()]

        if not points:
            return {"points": [], "summary": None}

        speeds = [p["speed"] or 0 for p in points]
        summary = {
            "total_km":    round(_calc_km(points), 2),
            "max_speed":   round(max(speeds), 1),
            "avg_speed":   round(sum(speeds)/len(speeds), 1) if speeds else 0,
            "point_count": len(points),
            "start_time":  points[0]["created_at"],
            "end_time":    points[-1]["created_at"],
            "stop_count":  _count_stops(points),
        }
        return {"points": points, "summary": summary}

    # ── GET /app/history/days ─────────────────────────────────
    @app.get("/app/history/days")
    def get_days(vehicle_id: str, days: int = 30, client_id: str = Query(...)):
        since = (datetime.utcnow() - timedelta(days=days)).isoformat()
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute("""
                SELECT DATE(created_at::timestamp) as day,
                       COUNT(*) as point_count,
                       MAX(speed) as max_speed,
                       MIN(created_at) as first_seen,
                       MAX(created_at) as last_seen
                FROM positions
                WHERE vehicle_id = %s AND created_at >= %s
                GROUP BY DATE(created_at::timestamp)
                ORDER BY day DESC
            """, (vehicle_id, since))
            rows = [dict(r) for r in cur.fetchall()]
        return {"days": [
            {
                "date": str(r["day"]),
                "point_count": r["point_count"],
                "max_speed": round(r["max_speed"] or 0, 1),
                "first_seen": r["first_seen"],
                "last_seen":  r["last_seen"],
            }
            for r in rows
        ]}

    # ── GET /app/reports/pdf ──────────────────────────────────
    @app.get("/app/reports/pdf")
    def report_pdf(
        vehicle_id: str,
        date_from: str = Query(...),
        date_to: str = Query(...),
        client_id: str = Query(...),
    ):
        from fastapi.responses import StreamingResponse
        import io
        data = _get_report_data(vehicle_id, date_from, date_to)
        pdf  = _make_pdf(data)
        fname = f"reporte_{data['vehicle']['name']}_{date_from}_{date_to}.pdf".replace(" ", "_")
        return StreamingResponse(io.BytesIO(pdf), media_type="application/pdf",
                                 headers={"Content-Disposition": f"attachment; filename={fname}"})

    # ── GET /app/reports/excel ────────────────────────────────
    @app.get("/app/reports/excel")
    def report_excel(
        vehicle_id: str,
        date_from: str = Query(...),
        date_to: str = Query(...),
        client_id: str = Query(...),
    ):
        from fastapi.responses import StreamingResponse
        import io
        data  = _get_report_data(vehicle_id, date_from, date_to)
        excel = _make_excel(data)
        fname = f"reporte_{data['vehicle']['name']}_{date_from}_{date_to}.xlsx".replace(" ", "_")
        return StreamingResponse(io.BytesIO(excel),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={fname}"})


# ── Helpers ───────────────────────────────────────────────────

def _get_report_data(vehicle_id, date_from, date_to):
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute("SELECT name, plate FROM vehicles WHERE id = %s", (vehicle_id,))
        vehicle = dict(cur.fetchone() or {"name": vehicle_id, "plate": ""})
        cur.execute("""
            SELECT DATE(created_at::timestamp) as day,
                   COUNT(*) as points,
                   MAX(speed) as max_speed,
                   AVG(speed) as avg_speed,
                   MIN(created_at) as start_time,
                   MAX(created_at) as end_time
            FROM positions
            WHERE vehicle_id = %s
              AND DATE(created_at::timestamp) BETWEEN %s::date AND %s::date
            GROUP BY DATE(created_at::timestamp)
            ORDER BY day ASC
        """, (vehicle_id, date_from, date_to))
        rows = [dict(r) for r in cur.fetchall()]

    processed, total_km = [], 0.0
    for r in rows:
        est_km = round((r["avg_speed"] or 0) * (r["points"] * 10 / 3600), 2)
        total_km += est_km
        processed.append({
            "date":    str(r["day"]),
            "start":   _fmt(r["start_time"]),
            "end":     _fmt(r["end_time"]),
            "minutes": _diff_min(r["start_time"], r["end_time"]),
            "max_speed": round(r["max_speed"] or 0, 1),
            "avg_speed": round(r["avg_speed"] or 0, 1),
            "est_km":  est_km,
        })
    return {"vehicle": vehicle, "date_from": date_from, "date_to": date_to,
            "rows": processed, "total_km": round(total_km, 2),
            "generated_at": datetime.utcnow().strftime("%d/%m/%Y %H:%M UTC")}


def _make_pdf(data):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    import io
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=2*cm, rightMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story = [
        Paragraph(f"GPS Control EC — Reporte de Actividad", ParagraphStyle("t", parent=styles["Heading1"], fontSize=15, textColor=colors.HexColor("#0d1117"))),
        Spacer(1, 0.3*cm),
        Paragraph(f"Vehículo: <b>{data['vehicle']['name']}</b>  |  Placa: {data['vehicle']['plate']}  |  {data['date_from']} → {data['date_to']}", styles["Normal"]),
        Paragraph(f"Generado: {data['generated_at']}  |  Total estimado: <b>{data['total_km']} km</b>", styles["Normal"]),
        Spacer(1, 0.6*cm),
    ]
    headers = ["Fecha", "Inicio", "Fin", "Min", "Vel máx", "Vel prom", "Km est."]
    td = [headers] + [[r["date"],r["start"],r["end"],str(r["minutes"]),str(r["max_speed"]),str(r["avg_speed"]),str(r["est_km"])] for r in data["rows"]]
    td.append(["TOTAL","","","","","",str(data["total_km"])])
    t = Table(td, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#0d1117")),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,-1),8),
        ("ALIGN",(0,0),(-1,-1),"CENTER"),
        ("ROWBACKGROUNDS",(0,1),(-1,-2),[colors.white,colors.HexColor("#f5f5f5")]),
        ("BACKGROUND",(0,-1),(-1,-1),colors.HexColor("#e8e8e8")),
        ("FONTNAME",(0,-1),(-1,-1),"Helvetica-Bold"),
        ("GRID",(0,0),(-1,-1),0.4,colors.lightgrey),
    ]))
    story.append(t)
    doc.build(story)
    return buf.getvalue()


def _make_excel(data):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import io
    wb = Workbook(); ws = wb.active; ws.title = "Reporte"
    ws.merge_cells("A1:G1")
    ws["A1"] = f"GPS Control EC — {data['vehicle']['name']} ({data['vehicle']['plate']})"
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="0d1117")
    ws["A1"].alignment = Alignment(horizontal="center")
    for col, h in enumerate(["Fecha","Inicio","Fin","Min","Vel máx","Vel prom","Km est."],1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="00b4d8")
        c.alignment = Alignment(horizontal="center")
    for i, r in enumerate(data["rows"]):
        for col, v in enumerate([r["date"],r["start"],r["end"],r["minutes"],r["max_speed"],r["avg_speed"],r["est_km"]],1):
            c = ws.cell(row=4+i, column=col, value=v)
            c.alignment = Alignment(horizontal="center")
            if i%2==0: c.fill = PatternFill("solid", fgColor="f0f0f0")
    ws.cell(row=4+len(data["rows"]),column=1,value="TOTAL").font=Font(bold=True)
    ws.cell(row=4+len(data["rows"]),column=7,value=data["total_km"]).font=Font(bold=True)
    for col,w in enumerate([12,9,9,8,10,10,9],1):
        ws.column_dimensions[ws.cell(row=1,column=col).column_letter].width=w
    buf=io.BytesIO(); wb.save(buf); return buf.getvalue()


def _haversine(lat1,lng1,lat2,lng2):
    R=6371; φ1,φ2=math.radians(lat1),math.radians(lat2)
    dφ=math.radians(lat2-lat1); dλ=math.radians(lng2-lng1)
    a=math.sin(dφ/2)**2+math.cos(φ1)*math.cos(φ2)*math.sin(dλ/2)**2
    return R*2*math.atan2(math.sqrt(a),math.sqrt(1-a))

def _calc_km(pts):
    t=0.0
    for i in range(1,len(pts)):
        if pts[i-1]["lat"] and pts[i]["lat"]:
            t+=_haversine(pts[i-1]["lat"],pts[i-1]["lng"],pts[i]["lat"],pts[i]["lng"])
    return t

def _count_stops(pts):
    s,in_s=0,False
    for p in pts:
        if (p["speed"] or 0)<2:
            if not in_s: s+=1; in_s=True
        else: in_s=False
    return s

def _fmt(dt_str):
    try: return datetime.fromisoformat(str(dt_str)).strftime("%H:%M")
    except: return "--"

def _diff_min(s,e):
    try: return int((datetime.fromisoformat(str(e))-datetime.fromisoformat(str(s))).total_seconds()/60)
    except: return 0